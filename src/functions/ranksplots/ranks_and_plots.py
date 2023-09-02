"""
This code is to generate the rank and plots of the model in python
"""

# from ast import Delete, Param
# from curses import raw
# from itertools import groupby
import os,io,sys,logging,math,openpyxl,string
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import statsmodels.api as sm
from pandas.core.base import DataError
from openpyxl.styles import Font,Border,Side,Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.copier import WorksheetCopy
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.drawing.image import Image
from openpyxl.utils.cell import coordinate_from_string
from scipy import stats

# Import funtion used in this code
from sys.path.insert(0,os.path.abspath(os.path.join(os.path.dirname(__file__),'..')))
from src.functions.univariate.predictive_power import univariate_predictive_power

stdout_handler=logging.StreamHandler(sys.stdout)
logging.basicConfig(level=logging.INFO, format='[%(levelname)s][%(asctime)s] - %(message)s', handlers=[stdout_handler])
logger=logging.getLogger('Ranks&plots logger')

class RanksPlots:
    def __init__(self,df,variables,target_variable,target_type,smvs=None,no_of_bins=20,fraction=1,weight_variable=None, data_dictionary=pd.DataFrame(),\
                  output_path=os.path.dirname(os.getcwd()),outputfilename='RanksandPlots.xlsx',raw=True):
        """
        Constructor
        """
        self._df=df
        self._variables=variables
        self._target_variable=target_variable
        self._target_type=target_type
        self._smvs=smvs
        self._no_of_bins=no_of_bins
        self._fraction=fraction
        self._weight_variable=weight_variable
        self._data_dictionary=data_dictionary
        self._output_path=output_path
        self._outputfilename=outputfilename
        self._raw=raw

        if self._target_type not in ['binary','continuous']:
            raise ValueError("target type must be either 'binary' or 'continuous' ")
        if not all(t in ('categorical','continuous') for t in set(self._variables.values())):
            raise ValueError("Variable type must be on of ('categorical','continuous')")
        
        self._metrics_table = {}   # key:value -> variablename:meanXperBin
        self._summary_table = {}   # key:value -> variablename:univariate
        self._scatter_plots = {}   # key:value -> variablename:matplotlib.figure.Figure
        self._correlations = {}    
        self.correlation = {}
        self.description = {}
    def run(self) -> None:
        """
        Generate Ranks and plots for self._variables
        
        return : None
        """
        if self._raw:
            logger.info("Application Started - Running for Raw variables")
        else:
            logger.info("Application started - Running for Transformed Variables")
        
        # Step 1 : Creates the Bins
        logger.info("creating the bins for the variables...")
        self._df=self.create_bins()
        logger.info("Bins have been created successfully")

        for variable,var_type in self._variables.items():
            # get any special missing values from smvs dictionaries
            if variable in self._smvs:
                smvs=self._smvs[variable]
            else:
                smvs=None

            logger.info(f"Processing variable {variable} of type {var_type}")
            # Step 2 Compute the means for every bins
            self._metrics_table[variable]=self.compute_metrics(self._df, variable)

            if self._target_type=='binary':
                # Step 3 Compute log odds
                logger.info("\t computing logodds")
                self._metrics_table[variable]= self.compute_log_odds(self._metrics_table[variable])
                logger.info("\t logodds computer successfully")

                self._correlations[variable]=self.correlation_for_bins(self._metrics_table[variable],variable,smvs)

            # Step 4 : Calculate univariate summary statistics  - Gini (binary) / R2 (continuous)
            # Get description and relationship from data dictionary

            logger.info("\t computing summary statistics...")
            self._summary_table[variable]=self.summary_statistics(self._df,variable,smvs)
            # Variable with no description available such as derive variable add description
            var_list=[]
            for var in self._data_dictionary['Variable']:
                var_list.append(var)
            if len(self._data_dictionary)>0:
                if variable in var_list:
                    self.description[variable]=self._data_dictionary.loc[self._data_dictionary['Variable']==variable,'Description'].values[0]
                else:
                    self.description[variable]="NA"
            else:
                self.description[variable]=""
            
            # Step 5 - Generate the scatter plots
            logger.info("\t Generating the scatter plot figure")
            self._scatter_plots[variable]=self.create_figure(self._metrics_table[variable],variable,var_type,smvs)
            plt.close(self._scatter_plots[variable]) # Avoid memory leak
            logger.info("\t Figure has been created successfully")

        # Step 6: Create the output
        logger.info("Creating the output file..")
        self.generate_output()
        logger.info(f"Output file has been created under {self._output_path}")
        logger.info("END of Application..")
    
    def create_bins(self):
        """
        Create the bins
        """
        df=self._df.copy(deep=True)

        for variable, var_type in self._variables.items():
            logger.info(f"\t Processing variable {variable} of type {var_type}")
            # get any special missing values from smvs dictionaries
            if variable in self._smvs[variable]:
                smvs=self._smvs[variable]
            else:
                smvs=None
            if var_type == 'continuous':
                if smvs:
                    x_vals=df[~df[variable].isin(smvs)].copy(deep=True)
                    x_smvs=df[df[variable].isin(smvs)].copy(deep=True)
                else:
                    x_vals = df.copy(deep=True)
                no_unique_vals=x_vals[variable].nunique()
                logger.info(f"\t Number of unique values:{no_unique_vals}")
                if no_unique_vals <=10:
                    x_vals[variable+'_binned']=x_vals[variable]
                else:
                    # Dynamically adjust the number of bins based on the unique values per variables
                    no_bins=self._no_of_bins
                    if no_bins>no_unique_vals -1:
                        no_bins=no_unique_vals-1
                    # Bin on non genuine or Special Missing values
                    x_vals[variable+'_binned']=pd.cut(x_vals[variable].rank(method='min'),no_bins,labels=False)
                # Crease Bins sequencially by 1
                x_vals[variable+'_binned']=x_vals[variable+'_binned'].rank(method='dense')
                # set bin as string
                x_vals[variable+'_binned']=x_vals[variable+'_binned'].astype('Int64')

                # Ensure missings are included as a bin
                x_vals[variable+'_binned']=np.where((x_vals[variable].isna()),'Missing',x_vals[variable+'_binned'])

                # if contains special missing values then ensure they are in one bin
                if smvs:
                    x_smvs[variable+'_binned']="SMVS: "+x_smvs[variable].astype('Int64').astype(str)
                    x_vals=pd.concat([x_vals,x_smvs], axis=0)
                df = x_vals.copy()
            elif var_type == 'categorical':
                # each unique value is a bin on its own
                df[variable+'_binned']=df[variable]
                if df[variable+'_binned'].nunique()>1000:
                    logger.warning(f"more than 1000 bins were created for variable {variable}")
        return df
    def compute_metrics(self,df,variable):
        """
        Compute mean(x), min(x), max(x), count per bin, %per bin
        """
        _df=df.copy(deep=True)
        def std(x):
            return np.std(x)
        def weighted_average(data):
            wt_avg=np.average(data.iloc[:,0],weights=data.iloc[:,1])
            return wt_avg
        def weighted_std(data):
            wt_avg=weighted_average(data)
            variance=np.average((data.iloc[:,0]-wt_avg)**2, weights=data.iloc[:,1])
            return math.sqrt(variance)
        #compute the mean, min, max and count for the variable per bin
        try:
            if self._weight_variable:
                #Below line of code is used only to get the data error for categorical feature
                _tmp=df.groupby(variable+'_binned')[variable].agg(['mean'])
                _df_mean=df.groupby(variable+'_binned')[[variable, self._weight_variable]].apply(weighted_average).reset_index(name='mean')
                _df=df.groupby(variable+'_binned')[variable].agg(['min','max']).reset_index()
                _df=_df_mean.merge(_df, on=[variable+'_binned'],how='outer')
            else:
                _df=df.groupby(variable+'_binned')[variable].agg(['min','max','count']).reset_index()
        except DataError: #Categorical feature
            if self._weight_variable:
                _df=df.groupby(variable+'_binned')[variable].agg(['min','max']).reset_index()
                _df['mean']=_df[variable+'_binned']
                _df=_df[[variable+'_binned','mean','min','max','count']]
        
        #Update (FN): Improvise code supplies by om
        if self._weight_variable:
                wt_count=df.groupby(variable+'_binned')[self._weight_variable].agg(['sum']).reset_index().rename(columns={'sum':'count'})
                _df=_df.merge(wt_count,on=[variable+'_binned'],how='left')
        else:
            n_missing=df[variable].isna().sum()
            if n_missing > 0:
                _df.loc[_df[variable+'_binned']=='Missing','count'] = df[variable].isna().sum()
            
        # Compute the percentage of frequency
        _df['countPercentage']=_df['count']/(_df['count'].sum())

        if self._weight_variable:
            # Compute the mean of the variable
            target_mean_df=df.groupby(variable+'_binned')[[self._target_variable,self._weight_variable]].apply(weighted_average).reset_index(name='targetVariableMean')
            _df=_df.merge(target_mean_df,on=[variable+'_binned'],how='left')
            # Compute the standard deviation of the target variable per bin
            target_std_df=df.groupby(variable+'_binned')[[self._target_variable,self._weight_variable]].apply(weighted_std).reset_index(name='targetVariableStd')
            _df=_df.merge(target_std_df, on=[variable+'_binned'],how='left')
        else:
            # Compute the mean of the target variable per bin
            _df['targetVariableMean']=df.groupby(variable+'_binned')[self._target_variable].agg(['mean']).reset_index()['mean']
            # Compute the standard variable of the target variable per bin
            _df['targetVariableStd']=df.groupby(variable+'_binned')[self._target_variable].agg(['std']).reset_index()['std']
        # Compute Sample err
        _df['standardError']=_df['targetVariableStd']/np.sqrt(_df['count'])
        # Compute lower and upper error
        # assumes gaussian distribution
        _df['lowerBound']=np.maximum(0,_df['targetVariableMean']-1.98*_df['targetVariableStd']/np.sqrt(_df['count']))
        _df['upperBound']=np.maximum(0,_df['targetVariableMean']+1.98*_df['targetVariableStd']/np.sqrt(_df['count']))
        # Remove strings from min/max columns
        mask=_df[['min','max']].applymap(lambda x: isinstance(x,(int,float)))
        _df[['min','max']]=_df[['min','max']].where(mask)
        return _df
    @staticmethod
    def compute_log_odds(df):
        """ Compute the log odds for the target variable"""
        _df=df.copy(deep=True)
        eps=1.e-15 #used to avoid divisions by 0
        _df['targetVariableMeanLogodds']=np.log((_df['targetVariableMean']+eps)/(1-_df['targetVariableMean']+eps))
        _df['standardErrorLogOddsLower']=np.log((_df['lowerBound']+eps)/(1-_df['lowerBound']+eps))
        _df['standardErrorLogOddsUpper']=np.log((_df['upperBound']+eps)/(1-_df['upperBound']+eps))
        # We need to compute those manually in order to properly plot error bars
        _df['YERR1']=_df['targetVariableMeanLogodds']-_df['standardErrorLogOddsLower']
        _df['YERR2']=_df['standardErrorLogOddsUpper']-_df['targetVariableMeanLogodds']
        #ToDo Need to handle an edge case
        # in the plot with many variables one of the bin has only single entry
        # therefore standard deviation =0
        return _df
    def create_figure(self,df,variable_name,var_type,smvs,border=True):
        """
        Create a scatter plot where:
        - x-axis -> mean(variable_name) per bin
        - y-axis -> mean(log odds)(binary)/mean( target value) (continuous)
        -  with a fitted line using statsmodels.nonparametric.lowess
        """
        _df=df.copy(deep=True)
        # Remove missing from plot data
        _df=_df[_df[variable_name+'_binned'].astype(str)!="Missing"]
        # remove smvs
        if smvs:
            _df=_df[~_df[variable_name+'_binned'].astype(str).isin(["SMV: "+str(int(smv)) for smv in smvs])]
        if border:
            fig=plt.figure(figsize=(8,6), linewidth=2, edgecolor="black")
        else:
            fig=plt.figure(figsize=(8,6))
        if self._target_type=='binary':
            # Add the scatter plot
            plt.scatter(_df['mean'],_df['targetVariableMeanLogodds'])

            # Add the error bar
            plt.errorbar(_df['mean'],_df['targetVariableMeanLogodds'],yerr=[_df['YERR1'].tolist(),_df['YERR2'].tolist()],fmt='o')
            plt.title(f"Ranks and Plots for predictor variable:{variable_name}")
            plt.xlabel(f"Mean({variable_name}) per Bin")
            plt.ylabel(f"Log Odds")
            if var_type != 'categorical':
                fitted_regression_results=self.apply_lowess_function(_df)
                plt.plot(_df['mean'], fitted_regression_results,'r--',label='lowess')
                plt.legend()
            return fig
        else:
            # Add the Scatter Plot
            plt.scatter(_df['mean'],_df['targetVariableMean'])
            plt.title(f"Ranks and Plots for predictor variable:{variable_name}")
            plt.xlabel(f"Mean({variable_name}) per Bin")
            plt.ylabel(f"Mean Target")
            if var_type != 'categorical':
                fitted_regression_results=self.apply_lowess_function(_df)
                plt.plot(_df['mean'], fitted_regression_results,'r--',label='lowess')
                plt.legend()
            return fig
    def apply_lowess_function(self,df):
        """ Apply a lowes function that outs smoothed estimates of x at the given y values from points (y,x)"""
        lowess=sm.nonparametric.lowess
        if self._target_type=='binary':
            try:
                result=lowess(df['targetVariableMeanLogodds'],df['mean'],frac=self._fraction)
            except ValueError:
                result=lowess(df['targetVariableMean'],list(range(0,df['mean'].nunique())),frac=self._fraction)
            return result[:,1]
        else:
            try:
                result=lowess(df['targetVariableMean'],df['mean'],frac=self._fraction)
            except ValueError:
                result=lowess(df['targetVariableMean'],list(range(0,df['mean'].nunique())),frac=self._fraction)
            return result[:,1]
    def summary_statistics(self,df,variable,smvs=None):
        """
        Calculate summary statistics -Gini/R2
        """
        _df=self._df[[variable,self._target_variable]+([self._weight_variable] if self._weight_variable else [])].copy()
        # Remove special missing values from plot data (if any)
        if smvs:
            _df=_df[~_df[variable].isin(smvs)]
        try:
            uni_results=univariate_predictive_power(_df,target_variable=self._target_variable,weight=self._weight_variable,bins=20,num_categories=20,num_transforms=False,
                                                    invalid_obs_cutoff=0.1,njobs=1,target_type=self._target_type)  #,sp_out
            if self._target_type=='binary':
                uni_result=uni_results[2]['Gini'][0]
            else:
                uni_result=uni_results[1]['rsq'][0]
        except:
            logger.warning(f"Unable to calculate univariate score for {variable}")
            uni_result=np.Nan
        return uni_result
    def correlation_for_bins(self,df,variable,smvs=None):
        """ CAlculate ZCorrelation"""
        _df=df.copy(deep=True)
        _df=_df[_df[variable+'_binned'] != "Missing"]
        if smvs:
            _df=_df[~_df[variable_name+'_binned'].astype(str).isin(["SMV: "+str(int(smv)) for smv in smvs])]
        try:
            pc=stats.pearsonr(_df['mean'],_df['targetVariableMeanLogodds'])
            correlation=np.where(pc[1]>0.1,"Ambiguous",np.where(pc[0]>0,"Positive","Negative"))
            corrrelation=correlation.item()
            return corrrelation
        except:
            print(f"Unable to calculate correlation for {variable}")
        
    def generate_output(self)->None:
        """ Generate the output file"""
        # Dictionaries of list containing Excel formatting parameters for each columns:
        col_parameters={'Rank':[Font(bold=True),Alignment(horizontal="left",vertical="center"),'General',150],
                        'Mean':[Font(bold=False),Alignment(horizontal="center",vertical="center"),'General',110],
                        'Min':[Font(bold=False),Alignment(horizontal="center",vertical="center"),'General',75],
                        'Max':[Font(bold=False),Alignment(horizontal="center",vertical="center"),'General',75],
                        'Volume':[Font(bold=False),Alignment(horizontal="center",vertical="center"),'#,##0',90],
                        'Volume%':[Font(bold=False),Alignment(horizontal="center",vertical="center"),'0.00%',90],
                        'target_Mean':[Font(bold=False),Alignment(horizontal="center",vertical="center"),'0.00%',108],
                        'log-odds':[Font(bold=False),Alignment(horizontal="center",vertical="center"),'0.00',84]}
        if self._target_type!='binary':
            col_parameters['target_Mean'][2]='#,##0.00'
        params=list(col_parameters.values())
        plot_length=sum([i[3] for i in params[:-1]])

        #BORDER STYLE
        thin_border=Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))
        # if raw is true then load an blank template workbook - should be saved in same folder as this script
        if self._raw:
            wb=openpyxl.load_workbook('resources/rank_and_plot_template.xlsx')
            start_col=5
            if self._target_type=="binary":
                template=wb['rnp_template_binary']
            else:
                template=wb['rnp_template_continuous']
        # Else add the existing template
        else:
            wb=openpyxl.load_workbook(os.path.join(self._output_path,self._outputfilename))
            if self._target_type=="binary":
                start_col=15
            else:
                start_col=14
        for variable,var_type in self._variables.items():
            logger.info(f"\t Writting table and figures for variable {variable}")
            # Rename output table headers
            rename_cols={variable+'_binned':'Rank',
                         'mean':'Mean',
                         'min':'Min',
                         'max':'Max',
                         'count':'Volume',
                         'countPercentage':'Volume%',
                         'targetVariableMean':'target Mean'} 
            if self._target_type=='binary':
                rename_cols['targetVariableMeanLogodds']='log-odds'
            _df = self._metrics_table[variable][[*rename_cols.keys()]].copy()
            _df.rename(rename_cols,axis=1,inplace=True)
            #Copy template /select sheet
            if self._raw:
                variable_ws=wb.create_sheet(variable)
                instance=WorksheetCopy(template,variable_ws)
                WorksheetCopy.copy_worksheet(instance)
            else:
                variable_ws=wb[variable]
                # if transformed already populated from previous run, delete data and plot
                if self._target_type=='binary':
                    delete_range='P3:W100'
                else:
                    delete_range='O3:U100'
                if len(variable_ws._images)==2:
                    del variable_ws._images[1]
                    for row in variable_ws[delete_range]:
                        for cell in row:
                            cell.value=None
                            cell.style='Normal'
            # Save the pandas data frame to the excel worksheet
            for r_idx,row in enumerate(dataframe_to_rows(_df,index=False),1):
                for c_idx,value in enumerate(row,1):
                    c=variable_ws.cell(row=r_idx+2,column=c_idx+start_col,value=value)
                    # Formatting
                    c.border=thin_border
                    if r_idx==1:
                        c.font=Font(bold=True)
                        if c_idx==1:
                            c.alignment=Alignment(horizontal='left',vertical='center')
                        else:
                            c.alignment=Alignment(horizontal='center',vertical='center')
                    else:
                        c.font=params[c_idx-1][0]
                        c.alignment=params[c_idx-1][1]
                        c.number_format=params[c_idx-1][2]
                    max_row=r_idx+2
            #Adjusted Column lenth
            col_length=len(col_parameters)
            if self._target_type != 'binary':
                col_length-=1
            for col in variable_ws.iter_rows(min_row=1,min_col=start_col+1, max_col=start_col+col_length,max_row=1):
                for i,cell in enumerate(col):
                    colw=coordinate_from_string(cell.coordinate)[0]
                    variable_ws.column_dimensions[colw].width=params[i][3]/7
            #Ensure that both plots are saved at the same row regardless of how many bins there are
            # - if more bins in raw then transformed, then offset transformed
            # - if more bins in transformed than raw , the offset raw
            if self._raw:
                plot_offset=0
            else:
                last_raw_row=[cell for cell in variable_ws['F'] if cell.value][-1].row
                plot_offset=max(last_raw_row-max_row,0)
                if last_raw_row < max_row:
                    raw_rnp=Image(io.BytesIO(variable_ws._images[0]._data()))
                    del variable_ws._images[0]
                    hw_ratio=raw_rnp.height/raw_rnp.width
                    raw_rnp.height=hw_ratio*plot_length
                    raw_rnp.width=plot_length
                    raw_rnp.anchor=variable_ws.cell(row=max_row+1,column=6).coordinate
                    variable_ws.add_image(raw_rnp)
            # Save and format the scatter plot
            imgdata=io.BytesIO()
            self._scatter_plots[variable].savefig(imgdata,format='png',edgecolor=self._scatter_plots[variable].get_edgecolor())
            img=Image(imgdata)
            hw_ratio=img.height/img.width
            img.height=hw_ratio*plot_length
            img.width=plot_length
            img.anchor=variable_ws.cell(row=max_row+2+plot_offset,column=start_col+1).coordinate
            variable_ws.add_image(img)

            #Population info table
            if self._raw:
                variable_ws['C2']=variable
                variable_ws['C12']=self._summary_table[variable]
                variable_ws['C3']=self.description[variable]
                variable_ws['C10'].value=self.correlation[variable]
                if self._target_type=='binary':
                    # remove missing from R2 Calculations
                    r2_df=_df.copy()
                    if var_type=='continuous':
                        r2_df=r2_df[r2_df['Rank'].astype(str) != "Missing"]
                        if variable in self._smvs:
                            smvs=self._smvs[variable]
                        else:
                            smvs=None
                        if smvs:
                            r2_df=r2_df[~r2_df['Rank'].astype(str).isin(["SMV: "+str(int(smv)) for smv in smvs])]
                        r2_df=sm.add_constant(r2_df, has_constant='add')
                        linear_reg=sm.OLS(r2_df[['log_odds']],r2_df[['const','Mean']], missing='drop').fit()
                    else:
                        r2_df=r2_df.sort_values(['log_odds']).reset_index(drop=True)
                        r2_df['r2_bin']=r2_df.index.astype(str).str.zfill(2)
                        r2_df=pd.get_dummies(r2_df,columns=['r2_bin'],drop_first=True)
                        linear_reg=sm.OLS(r2_df[['log_odds']],r2_df.loc[:,r2_df.columns.str.contains('r2_bin_')],missing='drop').fit()
                    variable_ws['C14']=linear_reg.rsquared
                    variable_ws['C14'].number_format='0.00%'
            else:
                variable_ws['C13']=self._summary_table[variable]
                variable_ws['C13'].number_format='0.00%'
                if self._target_type=='binary':
                    r2_df=self._metrics_table[variable].copy()
                    r2_df=sm.add_constant(r2_df, has_constant='add')
                    linear_reg=sm.OLS(r2_df[['targetVariableMeanLogodds']],r2_df[['const','mean']], missing='drop').fit()
                    variable_ws['C15']=linear_reg.rsquared
                    variable_ws['C15'].number_format='0.00%'
            # Sheet view settings
            variable_ws.sheet_view.zoomScale=70
            variable_ws.sheet_view.showGridLines=False
        #Delete template sheets
        if self._raw:
            del wb['rnp_template_binary']
            del wb['rnp_template_continuous']
        # Save worksheet
        wb.save(os.path.join(self._output_path,self._outputfilename))